from __future__ import annotations
from datetime import date
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func

from app.db.session import SessionLocal
# предположим названия моделей (переименуй при необходимости):
from app.db.models import Product, Unit  # твои модели
from app.db.models import StockDoc as Document, StockLine as DocItem  # или documents/doc_items
from datetime import date
from fastapi import APIRouter, Query, Response
from app.services.reports import build_incoming_day_xlsx

router = APIRouter(prefix="/api/reports", tags=["reports"])

@router.get("/incoming/day.xlsx")
def incoming_day_xlsx(
    date_: date = Query(..., alias="date"),
    location_id: int | None = Query(None)
):
    data = build_incoming_day_xlsx(date_, location_id)
    filename = f"Накладная_{date_.isoformat()}.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"',
    }
    return Response(content=data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

ORG_NAME = 'ДОО "СОЛНЫШКО" г.Чолпон-Ата'
VIA_NAME = ''

def _thin_border():
    thin = Side(style="thin", color="000000")
    return Border(top=thin, bottom=thin, left=thin, right=thin)

def _set_border_range(ws, cell_range):
    for row in ws[cell_range]:
        for c in row:
            c.border = _thin_border()

def _autosize(ws):
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            txt = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(txt))
        ws.column_dimensions[letter].width = min(max_len + 2, 42)

def _ru_month(d: date) -> str:
    months = [
        "января", "февраля", "марта", "апреля", "мая", "июня",
        "июля", "августа", "сентября", "октября", "ноября", "декабря",
    ]
    return months[d.month - 1]

def build_incoming_day_xlsx(report_date: date, location_id: Optional[int] = None) -> bytes:
    """
    Генерирует накладную за выбранный день из приходных документов (IN).
    Группировка: по товару и цене (чтобы не смешивать разные цены в один ряд).
    Единица берем из Product.unit (если qty хранится в базовой единице, делим на to_base_factor).
    """
    db = SessionLocal()
    try:
        # doc_type: адаптируй под свою модель/enum
        # если у тебя documents/doc_items — замени имена полей.
        q = (
            db.query(
                Product.name.label("name"),
                Unit.code.label("uom"),
                func.sum(
                    func.coalesce(
                        # если есть qty_base — конвертируем в удобные единицы
                        (DocItem.qty / 1.0),  # fallback, перепишем ниже
                        0
                    )
                ).label("qty"),
                DocItem.price.label("price"),
            )
            .join(Product, Product.id == DocItem.product_id)
            .join(Unit, Unit.id == Product.unit_id)
            .join(Document, Document.id == DocItem.doc_id)
            .filter(Document.doc_type.in_(("IN","IN_")))  # поддержим оба варианта
            .filter(func.date(Document.doc_date) == report_date)
        )
        # Если у тебя поля называются иначе (qty_base, to_base_factor и т.п.), вот более точный вариант:
        # qty в удобной единице = CASE WHEN DocItem.qty_base IS NOT NULL THEN DocItem.qty_base / NULLIF(Product.to_base_factor,0) ELSE DocItem.qty END
        qty_expr = func.coalesce(
            (DocItem.qty if hasattr(DocItem, "qty") else 0)  # заменится ниже
            , 0
        )
        if hasattr(DocItem, "qty_base") and hasattr(Product, "to_base_factor"):
            qty_expr = func.coalesce(DocItem.qty_base / func.nullif(Product.to_base_factor, 0), 0)
        q = q.with_entities(
            Product.name.label("name"),
            Unit.code.label("uom"),
            func.sum(qty_expr).label("qty"),
            DocItem.price.label("price"),
        )
        if hasattr(Document, "location_id") and location_id:
            q = q.filter(Document.location_id == location_id)

        q = q.group_by(Product.name, Unit.code, DocItem.price).order_by(Product.name)
        rows = q.all()

        # Собираем Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Накладная (день)"

        # Шапка (статическая — меняется только дата)
        ws.merge_cells("A1:G1"); ws["A1"] = ORG_NAME
        ws["A1"].font = Font(bold=True, size=12)
        ws["A1"].alignment = Alignment(horizontal="left")

        ws.merge_cells("A2:G2"); ws["A2"] = f"Через кого: {VIA_NAME}"
        ws["A2"].alignment = Alignment(horizontal="left")

        ws.merge_cells("A3:C3"); ws["A3"] = "НАКЛАДНАЯ №"
        ws["A3"].font = Font(bold=True, size=12)

        ws.merge_cells("D3:G3")
        ws["D3"] = f"От {report_date.day} {_ru_month(report_date)} {report_date.year} года"
        ws["D3"].alignment = Alignment(horizontal="left")

        ws.append([])

        # Заголовок таблицы
        headers = ["№", "Наименование", "Ед.изм", "Отпущено", "Цена за ед.", "Стоимость без НДС", "Примечание"]
        ws.append(headers)
        for c in range(1, len(headers)+1):
            cell = ws.cell(ws.max_row, c)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor="F2F2F2")

        start_data_row = ws.max_row + 1

        # Данные
        total_qty = 0
        total_amount = 0.0
        for i, r in enumerate(rows, start=1):
            amount = float(r.qty or 0) * float(r.price or 0)
            total_qty += float(r.qty or 0)
            total_amount += amount
            ws.append([i, r.name, r.uom, r.qty, r.price, amount, ""])

        # Форматирование числовых колонок и границы
        for r in range(start_data_row, ws.max_row + 1):
            ws.cell(r, 1).alignment = Alignment(horizontal="center")
            ws.cell(r, 3).alignment = Alignment(horizontal="center")
            ws.cell(r, 4).number_format = "#,##0.###"
            ws.cell(r, 5).number_format = '#,##0.00" сом"'
            ws.cell(r, 6).number_format = '#,##0.00" сом"'

        _set_border_range(ws, f"A{start_data_row-1}:G{ws.max_row}")

        # ИТОГО
        ws.append(["Итого", "", "", total_qty, "", total_amount, ""])
        tot_row = ws.max_row
        for c in range(1, 8):
            ws.cell(tot_row, c).font = Font(bold=True)
            ws.cell(tot_row, c).border = _thin_border()
        ws.cell(tot_row, 4).number_format = "#,##0.###"
        ws.cell(tot_row, 6).number_format = '#,##0.00" сом"'

        # Подписи
        ws.append([])
        ws.append(["Директор:", "", "", "", "", "", "________________________"])
        ws.append(["Зам.директор:", "", "", "", "", "", "________________________"])

        # Ширины колонок
        widths = [5, 40, 10, 14, 16, 22, 22]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        _autosize(ws)
        bio = BytesIO()
        wb.save(bio)
        return bio.getvalue()
    finally:
        db.close()

